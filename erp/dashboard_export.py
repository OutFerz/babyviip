"""
Exportación del resumen del panel: CSV, Excel, PDF, HTML.
"""

import csv
import html
import re
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone

# openpyxl y reportlab se importan dentro de cada export (lazy) para que el
# proyecto arranque aunque la imagen Docker no tenga aún esos paquetes instalados.


def _safe_slug(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or "reporte"


def _stamp() -> str:
    # Fecha local (amigable para nombres de archivo)
    return timezone.localdate().isoformat()


def _attach(response: HttpResponse, filename: str) -> HttpResponse:
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_dashboard_csv(ctx):
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(["Babyviip — Resumen panel (CSV)"])
    w.writerow(["Generado", timezone.now().isoformat()])
    w.writerow([])
    w.writerow(["Indicador", "Valor"])
    w.writerow(["Productos", ctx["total_productos"]])
    w.writerow(["Categorías", ctx["total_categorias"]])
    w.writerow(["Productos stock bajo", ctx["productos_stock_bajo"]])
    w.writerow(["Variantes agotadas (visibles)", ctx["variantes_agotadas"]])
    w.writerow([])
    w.writerow(["Top vendidos (producto)", "Unidades"])
    for row in ctx["top_vendidos"]:
        w.writerow([row["variante__producto__nombre"], row["total"]])
    w.writerow([])
    w.writerow(["Top favoritos (producto)", "Marcas"])
    for row in ctx["top_favoritos"]:
        w.writerow([row["producto__nombre"], row["total"]])
    w.writerow([])
    w.writerow(["Término buscado", "Veces"])
    for row in ctx["top_busquedas"]:
        w.writerow([row["termino"], row["total"]])
    w.writerow([])
    w.writerow(["Venta #", "Fecha", "Cliente", "Total"])
    for v in ctx["ventas_recientes"]:
        w.writerow(
            [
                v.id,
                v.fecha.isoformat() if v.fecha else "",
                v.cliente.nombre if v.cliente else "",
                str(v.total),
            ]
        )
    response = HttpResponse(
        "\ufeff" + buf.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    return _attach(response, f"babyviip_panel_resumen_{_stamp()}.csv")


def export_dashboard_xlsx(ctx):
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse(
            "Falta la dependencia openpyxl. En el contenedor: "
            "pip install openpyxl  (o docker compose build web)",
            status=503,
            content_type="text/plain; charset=utf-8",
        )
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "KPIs"
    ws0.append(["Indicador", "Valor"])
    ws0.append(["Productos", ctx["total_productos"]])
    ws0.append(["Categorías", ctx["total_categorias"]])
    ws0.append(["Productos stock bajo (< umbral)", ctx["productos_stock_bajo"]])
    ws0.append(["Variantes visibles agotadas", ctx["variantes_agotadas"]])

    ws1 = wb.create_sheet("Top vendidos")
    ws1.append(["Producto", "Unidades"])
    for row in ctx["top_vendidos"]:
        ws1.append([row["variante__producto__nombre"], row["total"]])

    ws2 = wb.create_sheet("Top favoritos")
    ws2.append(["Producto", "Favoritos"])
    for row in ctx["top_favoritos"]:
        ws2.append([row["producto__nombre"], row["total"]])

    ws3 = wb.create_sheet("Búsquedas")
    ws3.append(["Término", "Veces"])
    for row in ctx["top_busquedas"]:
        ws3.append([row["termino"], row["total"]])

    ws4 = wb.create_sheet("Ventas recientes")
    ws4.append(["ID", "Fecha", "Cliente", "Total"])
    for v in ctx["ventas_recientes"]:
        ws4.append(
            [
                v.id,
                v.fecha.replace(tzinfo=None) if v.fecha else None,
                v.cliente.nombre if v.cliente else "",
                float(v.total),
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return _attach(response, f"babyviip_panel_resumen_{_stamp()}.xlsx")


def export_dashboard_pdf(ctx):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas
    except ImportError:
        return HttpResponse(
            "Falta la dependencia reportlab. En el contenedor: "
            "pip install reportlab  (o docker compose build web)",
            status=503,
            content_type="text/plain; charset=utf-8",
        )
    buf = BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 2 * cm
    p.setFont("Helvetica-Bold", 14)
    p.drawString(2 * cm, y, "Babyviip — Resumen panel")
    y -= 0.6 * cm
    p.setFont("Helvetica", 9)
    p.drawString(
        2 * cm,
        y,
        f"Generado: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}",
    )
    y -= 1 * cm
    p.setFont("Helvetica", 10)
    for label, val in [
        ("Productos", ctx["total_productos"]),
        ("Categorias", ctx["total_categorias"]),
        ("Prod. stock bajo", ctx["productos_stock_bajo"]),
        ("Variantes agotadas (visibles)", ctx["variantes_agotadas"]),
    ]:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont("Helvetica", 10)
        p.drawString(2 * cm, y, f"{label}: {val}")
        y -= 0.7 * cm
    y -= 0.3 * cm
    p.setFont("Helvetica-Bold", 11)
    if y < 3 * cm:
        p.showPage()
        y = height - 2 * cm
    p.drawString(2 * cm, y, "Top vendidos")
    y -= 0.6 * cm
    p.setFont("Helvetica", 9)
    for row in ctx["top_vendidos"]:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont("Helvetica", 9)
        name = (row["variante__producto__nombre"] or "")[:80]
        p.drawString(2 * cm, y, f"{name} — {row['total']}")
        y -= 0.5 * cm
    y -= 0.3 * cm
    p.setFont("Helvetica-Bold", 11)
    if y < 3 * cm:
        p.showPage()
        y = height - 2 * cm
    p.drawString(2 * cm, y, "Ventas recientes")
    y -= 0.6 * cm
    p.setFont("Helvetica", 9)
    for v in ctx["ventas_recientes"]:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont("Helvetica", 9)
        line = f"#{v.id} {v.total} {v.fecha.strftime('%Y-%m-%d') if v.fecha else ''}"
        p.drawString(2 * cm, y, line[:100])
        y -= 0.5 * cm
    y -= 0.3 * cm
    p.setFont("Helvetica-Bold", 11)
    if y < 3 * cm:
        p.showPage()
        y = height - 2 * cm
    p.drawString(2 * cm, y, "Top favoritos")
    y -= 0.6 * cm
    p.setFont("Helvetica", 9)
    for row in ctx["top_favoritos"]:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont("Helvetica", 9)
        name = (row["producto__nombre"] or "")[:80]
        p.drawString(2 * cm, y, f"{name} — {row['total']}")
        y -= 0.5 * cm
    p.save()
    pdf = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    return _attach(response, f"babyviip_panel_resumen_{_stamp()}.pdf")


def export_dashboard_html(ctx):
    rows_v = "".join(
        f"<tr><td>{html.escape(str(r['variante__producto__nombre'] or ''))}</td>"
        f"<td>{r['total']}</td></tr>"
        for r in ctx["top_vendidos"]
    )
    rows_f = "".join(
        f"<tr><td>{html.escape(str(r['producto__nombre'] or ''))}</td>"
        f"<td>{r['total']}</td></tr>"
        for r in ctx["top_favoritos"]
    )
    rows_b = "".join(
        f"<tr><td>{html.escape(str(r['termino'] or ''))}</td><td>{r['total']}</td></tr>"
        for r in ctx["top_busquedas"]
    )
    rows_ventas = "".join(
        f"<tr><td>{v.id}</td><td>{html.escape(v.fecha.strftime('%d/%m/%Y %H:%M') if v.fecha else '')}</td>"
        f"<td>{html.escape(v.cliente.nombre if v.cliente else '—')}</td>"
        f"<td>{v.total}</td></tr>"
        for v in ctx["ventas_recientes"]
    )
    gen = html.escape(timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M"))
    body = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><title>Babyviip — Resumen panel</title>
<meta name="viewport" content="width=device-width, initial-scale=1" />
<style>
body{{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:1.5rem;color:#0f172a;}}
.card{{max-width:920px;margin:0 auto;border:1px solid #e2e8f0;border-radius:14px;padding:1.25rem 1.25rem 1rem;}}
.brand{{display:flex;justify-content:space-between;gap:1rem;flex-wrap:wrap;align-items:baseline;}}
.brand h1{{margin:0;color:#0277bd;font-size:1.35rem;}}
.meta{{margin:0;color:#475569;font-size:.95rem;}}
h2{{margin:1.1rem 0 .5rem;font-size:1.05rem;}}
table{{border-collapse:collapse;width:100%;}}
th,td{{border:1px solid #e2e8f0;padding:8px 10px;text-align:left;vertical-align:top;}}
th{{background:#f8fafc;}}
footer{{margin-top:1rem;color:#64748b;font-size:.9rem;}}
</style></head><body>
<div class="card">
<div class="brand"><h1>Babyviip — Resumen panel</h1><p class="meta">Generado: {gen}</p></div>
<h2>KPIs</h2>
<table><tr><th>Indicador</th><th>Valor</th></tr>
<tr><td>Productos</td><td>{ctx['total_productos']}</td></tr>
<tr><td>Categorías</td><td>{ctx['total_categorias']}</td></tr>
<tr><td>Productos stock bajo</td><td>{ctx['productos_stock_bajo']}</td></tr>
<tr><td>Variantes agotadas</td><td>{ctx['variantes_agotadas']}</td></tr>
</table>
<h2>Top vendidos</h2>
<table><tr><th>Producto</th><th>Unidades</th></tr>{rows_v}</table>
<h2>Top favoritos</h2>
<table><tr><th>Producto</th><th>Favoritos</th></tr>{rows_f}</table>
<h2>Búsquedas</h2>
<table><tr><th>Término</th><th>Veces</th></tr>{rows_b}</table>
<h2>Ventas recientes</h2>
<table><tr><th>#</th><th>Fecha</th><th>Cliente</th><th>Total</th></tr>{rows_ventas}</table>
</div>
<footer>Exportado desde el panel operativo Babyviip.</footer>
</body></html>"""
    response = HttpResponse(body, content_type="text/html; charset=utf-8")
    return _attach(response, f"babyviip_panel_resumen_{_stamp()}.html")


# --- Exportación por gráfico (mismos datos que Chart.js: labels + values) ---

CHART_META = {
    "kpis": ("chart_kpis", "Indicadores_clave", "Indicadores clave"),
    "vendidos": ("chart_vendidos", "Mas_vendidos", "Más vendidos (unidades)"),
    "favoritos": ("chart_favoritos", "Mas_favoritos", "Más favoritos"),
}


def _chart_data(ctx, grafico: str):
    if grafico not in CHART_META:
        return None
    key, _, _ = CHART_META[grafico]
    block = ctx.get(key) or {}
    labels = list(block.get("labels") or [])
    values = list(block.get("values") or [])
    return labels, values


def export_chart_csv(ctx, grafico: str):
    meta = CHART_META.get(grafico)
    if not meta:
        return None
    _, slug, title = meta
    slug = _safe_slug(slug)
    labels, values = _chart_data(ctx, grafico)
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow([f"Babyviip — {title}"])
    w.writerow(["Generado", timezone.now().isoformat()])
    w.writerow([])
    w.writerow(["Etiqueta", "Valor"])
    for i, lab in enumerate(labels):
        val = values[i] if i < len(values) else ""
        w.writerow([lab, val])
    response = HttpResponse(
        "\ufeff" + buf.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    return _attach(response, f"babyviip_grafico_{slug}_{_stamp()}.csv")


def export_chart_xlsx(ctx, grafico: str):
    meta = CHART_META.get(grafico)
    if not meta:
        return None
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse(
            "Falta openpyxl. pip install openpyxl",
            status=503,
            content_type="text/plain; charset=utf-8",
        )
    _, slug, title = meta
    slug = _safe_slug(slug)
    labels, values = _chart_data(ctx, grafico)
    wb = Workbook()
    ws = wb.active
    ws.title = slug[:31]
    ws.append([title])
    ws.append(["Generado", timezone.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    ws.append(["Etiqueta", "Valor"])
    for i, lab in enumerate(labels):
        val = values[i] if i < len(values) else ""
        ws.append([lab, val])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return _attach(response, f"babyviip_grafico_{slug}_{_stamp()}.xlsx")


def export_chart_pdf(ctx, grafico: str):
    meta = CHART_META.get(grafico)
    if not meta:
        return None
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas
    except ImportError:
        return HttpResponse(
            "Falta reportlab. pip install reportlab",
            status=503,
            content_type="text/plain; charset=utf-8",
        )
    _, slug, title = meta
    slug = _safe_slug(slug)
    labels, values = _chart_data(ctx, grafico)
    buf = BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    height = A4[1]
    y = height - 2 * cm
    p.setFont("Helvetica-Bold", 12)
    p.drawString(2 * cm, y, f"Babyviip — {title[:70]}")
    y -= 0.9 * cm
    p.setFont("Helvetica", 9)
    p.drawString(
        2 * cm,
        y,
        f"Generado: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}",
    )
    y -= 0.8 * cm
    for i, lab in enumerate(labels):
        val = values[i] if i < len(values) else ""
        line = f"{str(lab)[:55]}  |  {val}"
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont("Helvetica", 9)
        p.drawString(2 * cm, y, line[:120])
        y -= 0.45 * cm
    p.save()
    pdf = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    return _attach(response, f"babyviip_grafico_{slug}_{_stamp()}.pdf")


def export_chart_html(ctx, grafico: str):
    meta = CHART_META.get(grafico)
    if not meta:
        return None
    _, slug, title = meta
    slug = _safe_slug(slug)
    labels, values = _chart_data(ctx, grafico)
    gen = html.escape(timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M"))
    rows = "".join(
        f"<tr><td>{html.escape(str(lab))}</td><td>{html.escape(str(values[i] if i < len(values) else ''))}</td></tr>"
        for i, lab in enumerate(labels)
    )
    body = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><title>Babyviip — {html.escape(title)}</title>
<meta name="viewport" content="width=device-width, initial-scale=1" />
<style>
body{{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:1.5rem;color:#0f172a;}}
.card{{max-width:920px;margin:0 auto;border:1px solid #e2e8f0;border-radius:14px;padding:1.25rem 1.25rem 1rem;}}
.brand{{display:flex;justify-content:space-between;gap:1rem;flex-wrap:wrap;align-items:baseline;}}
.brand h1{{margin:0;color:#0277bd;font-size:1.25rem;}}
.meta{{margin:0;color:#475569;font-size:.95rem;}}
table{{border-collapse:collapse;width:100%;margin-top:.75rem;}}
th,td{{border:1px solid #e2e8f0;padding:8px 10px;text-align:left;vertical-align:top;}}
th{{background:#f8fafc;}}
footer{{margin-top:1rem;color:#64748b;font-size:.9rem;}}
</style></head><body>
<div class="card">
<div class="brand"><h1>Babyviip — {html.escape(title)}</h1><p class="meta">Generado: {gen}</p></div>
<table><tr><th>Etiqueta</th><th>Valor</th></tr>{rows}</table>
</div>
<footer>Exportado desde el panel operativo Babyviip.</footer>
</body></html>"""
    response = HttpResponse(body, content_type="text/html; charset=utf-8")
    return _attach(response, f"babyviip_grafico_{slug}_{_stamp()}.html")


# --- Exportación por tabla (bloques de listado del dashboard) ---

TABLA_META = {
    "busquedas": ("Terminos_buscados", "Términos más buscados"),
    "ventas": ("Ventas_recientes", "Ventas recientes"),
    "ranking_vendidos": ("Ranking_unidades_vendidas", "Ranking · unidades vendidas"),
    "ranking_favoritos": ("Ranking_favoritos", "Ranking · favoritos"),
}


def export_tabla_csv(ctx, tabla: str):
    meta = TABLA_META.get(tabla)
    if not meta:
        return None
    slug, title = meta
    slug = _safe_slug(slug)
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow([f"Babyviip — {title}"])
    w.writerow(["Generado", timezone.now().isoformat()])
    w.writerow([])
    if tabla == "busquedas":
        w.writerow(["Término", "Veces"])
        for row in ctx["top_busquedas"]:
            w.writerow([row["termino"], row["total"]])
    elif tabla == "ventas":
        w.writerow(["#", "Fecha ISO", "Cliente", "Total"])
        for v in ctx["ventas_recientes"]:
            w.writerow(
                [
                    v.id,
                    v.fecha.isoformat() if v.fecha else "",
                    v.cliente.nombre if v.cliente else "",
                    str(v.total),
                ]
            )
    elif tabla == "ranking_vendidos":
        w.writerow(["Producto", "Unidades"])
        for row in ctx["top_vendidos"]:
            w.writerow([row["variante__producto__nombre"], row["total"]])
    elif tabla == "ranking_favoritos":
        w.writerow(["Producto", "Favoritos"])
        for row in ctx["top_favoritos"]:
            w.writerow([row["producto__nombre"], row["total"]])
    response = HttpResponse(
        "\ufeff" + buf.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    return _attach(response, f"babyviip_tabla_{slug}_{_stamp()}.csv")


def export_tabla_xlsx(ctx, tabla: str):
    meta = TABLA_META.get(tabla)
    if not meta:
        return None
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse(
            "Falta openpyxl.",
            status=503,
            content_type="text/plain; charset=utf-8",
        )
    slug, title = meta
    slug = _safe_slug(slug)
    wb = Workbook()
    ws = wb.active
    ws.title = slug[:31]
    ws.append([title])
    ws.append([])
    if tabla == "busquedas":
        ws.append(["Término", "Veces"])
        for row in ctx["top_busquedas"]:
            ws.append([row["termino"], row["total"]])
    elif tabla == "ventas":
        ws.append(["ID", "Fecha", "Cliente", "Total"])
        for v in ctx["ventas_recientes"]:
            ws.append(
                [
                    v.id,
                    v.fecha.replace(tzinfo=None) if v.fecha else None,
                    v.cliente.nombre if v.cliente else "",
                    float(v.total),
                ]
            )
    elif tabla == "ranking_vendidos":
        ws.append(["Producto", "Unidades"])
        for row in ctx["top_vendidos"]:
            ws.append([row["variante__producto__nombre"], row["total"]])
    elif tabla == "ranking_favoritos":
        ws.append(["Producto", "Favoritos"])
        for row in ctx["top_favoritos"]:
            ws.append([row["producto__nombre"], row["total"]])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return _attach(response, f"babyviip_tabla_{slug}_{_stamp()}.xlsx")


def export_tabla_pdf(ctx, tabla: str):
    meta = TABLA_META.get(tabla)
    if not meta:
        return None
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas
    except ImportError:
        return HttpResponse(
            "Falta reportlab.",
            status=503,
            content_type="text/plain; charset=utf-8",
        )
    slug, title = meta
    slug = _safe_slug(slug)
    buf = BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    height = A4[1]
    y = height - 2 * cm
    p.setFont("Helvetica-Bold", 11)
    p.drawString(2 * cm, y, f"Babyviip — {title[:75]}")
    y -= 0.8 * cm
    p.setFont("Helvetica", 8)
    p.drawString(
        2 * cm,
        y,
        f"Generado: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}",
    )
    y -= 0.6 * cm
    if tabla == "busquedas":
        for row in ctx["top_busquedas"]:
            if y < 2 * cm:
                p.showPage()
                y = height - 2 * cm
            p.drawString(2 * cm, y, f"{str(row['termino'])[:60]} — {row['total']}")
            y -= 0.42 * cm
    elif tabla == "ventas":
        for v in ctx["ventas_recientes"]:
            if y < 2 * cm:
                p.showPage()
                y = height - 2 * cm
            line = f"#{v.id}  {v.total}  {v.fecha.strftime('%Y-%m-%d') if v.fecha else ''}"
            p.drawString(2 * cm, y, line[:100])
            y -= 0.42 * cm
    elif tabla == "ranking_vendidos":
        for row in ctx["top_vendidos"]:
            if y < 2 * cm:
                p.showPage()
                y = height - 2 * cm
            p.drawString(
                2 * cm,
                y,
                f"{str(row['variante__producto__nombre'])[:65]} — {row['total']}",
            )
            y -= 0.42 * cm
    elif tabla == "ranking_favoritos":
        for row in ctx["top_favoritos"]:
            if y < 2 * cm:
                p.showPage()
                y = height - 2 * cm
            p.drawString(
                2 * cm, y, f"{str(row['producto__nombre'])[:65]} — {row['total']}"
            )
            y -= 0.42 * cm
    p.save()
    pdf = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    return _attach(response, f"babyviip_tabla_{slug}_{_stamp()}.pdf")


def export_tabla_html(ctx, tabla: str):
    meta = TABLA_META.get(tabla)
    if not meta:
        return None
    slug, title = meta
    slug = _safe_slug(slug)
    gen = html.escape(timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M"))
    if tabla == "busquedas":
        rows = "".join(
            f"<tr><td>{html.escape(str(r['termino'] or ''))}</td><td>{r['total']}</td></tr>"
            for r in ctx["top_busquedas"]
        )
        thead = "<tr><th>Término</th><th>Veces</th></tr>"
    elif tabla == "ventas":
        rows = "".join(
            f"<tr><td>{v.id}</td><td>{html.escape(v.fecha.strftime('%d/%m/%Y %H:%M') if v.fecha else '')}</td>"
            f"<td>{html.escape(v.cliente.nombre if v.cliente else '—')}</td><td>{v.total}</td></tr>"
            for v in ctx["ventas_recientes"]
        )
        thead = "<tr><th>#</th><th>Fecha</th><th>Cliente</th><th>Total</th></tr>"
    elif tabla == "ranking_vendidos":
        rows = "".join(
            f"<tr><td>{html.escape(str(r['variante__producto__nombre'] or ''))}</td>"
            f"<td>{r['total']}</td></tr>"
            for r in ctx["top_vendidos"]
        )
        thead = "<tr><th>Producto</th><th>Unidades</th></tr>"
    elif tabla == "ranking_favoritos":
        rows = "".join(
            f"<tr><td>{html.escape(str(r['producto__nombre'] or ''))}</td>"
            f"<td>{r['total']}</td></tr>"
            for r in ctx["top_favoritos"]
        )
        thead = "<tr><th>Producto</th><th>Favoritos</th></tr>"
    else:
        return None
    body = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><title>Babyviip — {html.escape(title)}</title>
<meta name="viewport" content="width=device-width, initial-scale=1" />
<style>
body{{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:1.5rem;color:#0f172a;}}
.card{{max-width:1024px;margin:0 auto;border:1px solid #e2e8f0;border-radius:14px;padding:1.25rem 1.25rem 1rem;}}
.brand{{display:flex;justify-content:space-between;gap:1rem;flex-wrap:wrap;align-items:baseline;}}
.brand h1{{margin:0;color:#0277bd;font-size:1.25rem;}}
.meta{{margin:0;color:#475569;font-size:.95rem;}}
table{{border-collapse:collapse;width:100%;margin-top:.75rem;}}
th,td{{border:1px solid #e2e8f0;padding:8px 10px;text-align:left;vertical-align:top;}}
th{{background:#f8fafc;}}
footer{{margin-top:1rem;color:#64748b;font-size:.9rem;}}
</style></head><body>
<div class="card">
<div class="brand"><h1>Babyviip — {html.escape(title)}</h1><p class="meta">Generado: {gen}</p></div>
<table>{thead}{rows}</table>
</div>
<footer>Exportado desde el panel operativo Babyviip.</footer>
</body></html>"""
    response = HttpResponse(body, content_type="text/html; charset=utf-8")
    return _attach(response, f"babyviip_tabla_{slug}_{_stamp()}.html")
